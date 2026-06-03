import argparse
import datetime as dt
import os
import re
import sqlite3
import sys

import openpyxl
from openpyxl.utils.datetime import from_excel


KNOWN_SHEET_NAMES = {
    "充值记录", "学生查询", "费用标准", "学生单价表", "学生费用明细", "教师薪资汇总",
    "教师副本", "课程拆分", "Claude Log",
}


def find_total_sheet(wb):
    for name in wb.sheetnames:
        if re.fullmatch(r"\d{1,2}月总表", name):
            return name
    if "4月总表" in wb.sheetnames:
        return "4月总表"
    raise RuntimeError("未找到月度总表（如 2月总表、3月总表、4月总表）")


def month_key_from_filename(path):
    match = re.search(r"(\d{4})年(\d{1,2})月", os.path.basename(path))
    if not match:
        return ""
    return f"{match.group(1)}-{match.group(2).zfill(2)}-01"


def month_key_from_sheet_name(name, year=None):
    match = re.fullmatch(r"(\d{1,2})月总表", text(name))
    if not match:
        return ""
    return f"{year or dt.date.today().year}-{match.group(1).zfill(2)}-01"


def text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def number(value):
    if value in (None, ""):
        return 0
    try:
        return float(value)
    except Exception:
        return 0


def optional_number(value):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        return float(value)
    except Exception:
        return None


def iso_date(value):
    if value in (None, ""):
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date().isoformat()
        except Exception:
            return ""
    raw = text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return dt.datetime.strptime(raw[:10], fmt).date().isoformat()
        except Exception:
            pass
    return raw


def month_key_from_date(value):
    raw = iso_date(value)
    match = re.fullmatch(r"(\d{4})-(\d{2})-\d{2}", raw)
    if not match:
        return ""
    return f"{match.group(1)}-{match.group(2)}-01"


def infer_month_key(workbook_path, wb, total_sheet_name, explicit_month):
    if explicit_month:
        return iso_date(explicit_month)
    q1_month = iso_date(wb[total_sheet_name]["Q1"].value)
    if q1_month:
        return q1_month
    file_month = month_key_from_filename(workbook_path)
    if file_month:
        return file_month
    sheet_month = month_key_from_sheet_name(total_sheet_name)
    if sheet_month:
        return sheet_month
    return "2026-04-01"


def split_students(value):
    return [part.strip() for part in re.split(r"[、,，;；]", text(value)) if part.strip()]


def salary_preserve_key(date, teacher, time_slot, classroom, grade, subject, student_names):
    students = sorted({
        re.sub(r"\s+", "", name)
        for name in split_students(student_names)
        if re.sub(r"\s+", "", name)
    })
    return (
        text(date),
        text(teacher),
        text(time_slot),
        text(classroom),
        text(grade),
        text(subject),
        "、".join(students),
    )


def valid_teacher_name(value):
    name = text(value)
    if not name or name == "合计" or name.startswith("#"):
        return ""
    if name in KNOWN_SHEET_NAMES or re.fullmatch(r"\d+月.*", name) or re.fullmatch(r".*学生费用汇总", name):
        return ""
    if re.fullmatch(r"\d{1,2}\.\d{1,2}[-－~～]\d{1,2}\.\d{1,2}", name):
        return ""
    return name


def parse_teacher_order(value):
    return [valid_teacher_name(name) for name in re.split(r"[、,，;；]", text(value)) if valid_teacher_name(name)]


def derive_status(lesson_status, course_status):
    lesson_status = text(lesson_status)
    course_status = text(course_status)
    if lesson_status == "试课":
        return "试课"
    if lesson_status == "考试":
        return "考试"
    if lesson_status == "上课（未缴费）":
        return "未缴费"
    if lesson_status == "请假" or course_status == "请假":
        return "请假"
    if course_status == "已上":
        return "已上"
    return "待上"


def ensure_schema_exists(conn):
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='lessons'"
    ).fetchone()
    if not row:
        raise RuntimeError("数据库未初始化。请先在 liming-course-system 目录运行 npm run init。")
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='teacher_adjustments_monthly'"
    ).fetchone()
    if not row:
        raise RuntimeError("数据库结构需要升级。请先运行 npm run init 以创建 teacher_adjustments_monthly。")
    lesson_columns = {
        row[1] for row in conn.execute("PRAGMA table_info(lessons)").fetchall()
    }
    required_columns = {"teacher_salary_source", "teacher_salary_rule_id"}
    if not required_columns.issubset(lesson_columns):
        raise RuntimeError("数据库结构需要升级。请先运行 npm run init 以创建教师薪资来源字段。")


def upsert_student(conn, name, grade):
    if not name:
        return
    conn.execute(
        """
        INSERT INTO students(name, grade) VALUES (?, ?)
        ON CONFLICT(name) DO UPDATE SET
          grade = COALESCE(NULLIF(excluded.grade, ''), students.grade)
        """,
        (name, grade or ""),
    )


def upsert_teacher(conn, name):
    if not name:
        return
    conn.execute("INSERT OR IGNORE INTO teachers(name) VALUES (?)", (name,))


def lesson_rows(ws, fallback_month_key):
    rows = []
    for row_idx in range(3, ws.max_row + 1):
        teacher = text(ws.cell(row_idx, 1).value)
        date = iso_date(ws.cell(row_idx, 2).value)
        lesson_status = text(ws.cell(row_idx, 3).value)
        time_slot = text(ws.cell(row_idx, 5).value)
        classroom = text(ws.cell(row_idx, 6).value)
        grade = text(ws.cell(row_idx, 7).value)
        subject = text(ws.cell(row_idx, 8).value)
        student_names = text(ws.cell(row_idx, 9).value)
        notes = text(ws.cell(row_idx, 10).value)
        course_status = text(ws.cell(row_idx, 11).value)
        teacher_salary = optional_number(ws.cell(row_idx, 12).value)
        if not any([teacher, date, lesson_status, time_slot, classroom, grade, subject, student_names, notes, course_status, teacher_salary]):
            continue
        if not date:
            continue
        row_month_key = month_key_from_date(date) or fallback_month_key
        rows.append({
            "row_idx": row_idx,
            "teacher": teacher,
            "date": date,
            "lesson_status": lesson_status,
            "time_slot": time_slot,
            "classroom": classroom,
            "grade": grade,
            "subject": subject,
            "student_names": student_names,
            "notes": notes,
            "course_status": course_status,
            "teacher_salary": teacher_salary,
            "teacher_salary_present": teacher_salary is not None,
            "month_key": row_month_key,
            "status": derive_status(lesson_status, course_status),
        })
    return rows


def import_lessons(conn, wb, total_sheet_name, month_key, replace):
    ws = wb[total_sheet_name]
    rows = lesson_rows(ws, month_key)
    existing_by_key = {}
    existing_by_salary_preserve_key = {}
    for existing in conn.execute(
        """
        SELECT date, teacher_name, time_slot, classroom, grade, subject, student_names, teacher_salary,
               teacher_salary_source, teacher_salary_rule_id
        FROM lessons
        WHERE month_key = ?
        """,
        (month_key,),
    ).fetchall():
        key = (text(existing[0]), text(existing[1]), text(existing[2]))
        existing_by_key.setdefault(key, []).append(existing)
        exact_key = salary_preserve_key(
            existing[0], existing[1], existing[2], existing[3],
            existing[4], existing[5], existing[6]
        )
        existing_by_salary_preserve_key.setdefault(exact_key, []).append(existing)
    if replace:
        pairs = sorted({(row["date"], row["month_key"]) for row in rows if row["date"]})
        if pairs:
            conn.executemany("DELETE FROM lessons WHERE date = ? AND month_key = ?", pairs)
        else:
            conn.execute("DELETE FROM lessons WHERE month_key = ?", (month_key,))
    inserted = 0
    for row in rows:
        upsert_teacher(conn, row["teacher"])
        for name in split_students(row["student_names"]):
            upsert_student(conn, name, row["grade"])
        teacher_salary = row["teacher_salary"]
        teacher_salary_source = "import" if row["teacher_salary_present"] else "empty"
        teacher_salary_rule_id = None
        if not row["teacher_salary_present"]:
            key = (text(row["date"]), text(row["teacher"]), text(row["time_slot"]))
            exact_key = salary_preserve_key(
                row["date"], row["teacher"], row["time_slot"], row["classroom"],
                row["grade"], row["subject"], row["student_names"]
            )
            exact_matches = existing_by_salary_preserve_key.get(exact_key, [])
            matches = exact_matches if len(exact_matches) == 1 else existing_by_key.get(key, [])
            existing = matches[0] if len(matches) == 1 else None
            if existing is not None and existing[7] is not None:
                teacher_salary = existing[7]
                teacher_salary_source = text(existing[8]) or "legacy"
                teacher_salary_rule_id = existing[9]
        conn.execute(
            """
            INSERT INTO lessons(
              teacher_name, date, lesson_status, time_slot, classroom, grade, subject,
              student_names, notes, course_status, status, teacher_salary, teacher_salary_source,
              teacher_salary_rule_id, month_key, sort_order
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row["teacher"],
                row["date"],
                row["lesson_status"] or "上课",
                row["time_slot"],
                row["classroom"],
                row["grade"],
                row["subject"],
                row["student_names"],
                row["notes"],
                row["course_status"] or "未上",
                row["status"],
                teacher_salary,
                teacher_salary_source,
                teacher_salary_rule_id,
                row["month_key"],
                row["row_idx"],
            ),
        )
        inserted += 1
    return inserted, rows


def backup_database(conn, db_path):
    backup_dir = os.path.join(os.path.dirname(db_path), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d%H%M%S")
    backup_path = os.path.join(backup_dir, f"pre_import_{stamp}.sqlite")
    try:
        backup_conn = sqlite3.connect(backup_path)
        conn.backup(backup_conn)
    except sqlite3.Error as exc:
        print(f"备份跳过：{exc}", file=sys.stderr)
        return ""
    finally:
        if "backup_conn" in locals():
            backup_conn.close()
    return backup_path


def import_recharges(conn, wb, month_key, replace=True):
    if "充值记录" not in wb.sheetnames:
        return 0
    ws = wb["充值记录"]
    if replace:
        conn.execute("DELETE FROM recharge_records WHERE month_key = ?", (month_key,))
    count = 0
    for row_idx in range(3, ws.max_row + 1):
        student_name = text(ws.cell(row_idx, 1).value)
        if not student_name:
            continue
        payload = (
            student_name,
            text(ws.cell(row_idx, 2).value),
            number(ws.cell(row_idx, 3).value),
            number(ws.cell(row_idx, 4).value),
            number(ws.cell(row_idx, 5).value),
            number(ws.cell(row_idx, 6).value),
            iso_date(ws.cell(row_idx, 7).value),
            text(ws.cell(row_idx, 8).value),
            month_key,
        )
        conn.execute(
            """
            INSERT INTO recharge_records(
              student_name, grade, prev_actual, prev_gift, cur_recharge, cur_gift,
              recharge_date, notes, source, month_key
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'source-workbook', ?)
            ON CONFLICT(student_name, month_key) DO UPDATE SET
              grade = COALESCE(NULLIF(excluded.grade, ''), recharge_records.grade),
              prev_actual = excluded.prev_actual,
              prev_gift = excluded.prev_gift,
              cur_recharge = excluded.cur_recharge,
              cur_gift = excluded.cur_gift,
              recharge_date = excluded.recharge_date,
              notes = excluded.notes,
              source = excluded.source
            """,
            payload,
        )
        upsert_student(conn, student_name, payload[1])
        count += 1
    return count


def import_student_pricing(conn, wb):
    if "学生单价表" not in wb.sheetnames:
        return 0
    ws = wb["学生单价表"]
    count = 0
    for row_idx in range(3, ws.max_row + 1):
        student_name = text(ws.cell(row_idx, 1).value)
        subject = text(ws.cell(row_idx, 2).value)
        helper = text(ws.cell(row_idx, 8).value)
        if (not student_name or not subject) and "|" in helper:
          parts = helper.split("|", 1)
          student_name = student_name or parts[0].strip()
          subject = subject or parts[1].strip()
        custom_price = ws.cell(row_idx, 3).value
        if not student_name or not subject or custom_price in (None, ""):
            continue
        conn.execute(
            """
            INSERT INTO student_pricing(student_name, subject, custom_price, notes)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(student_name, subject) DO UPDATE SET
              custom_price = excluded.custom_price,
              notes = excluded.notes
            """,
            (student_name, subject, number(custom_price), text(ws.cell(row_idx, 6).value)),
        )
        count += 1
    return count


def import_pricing_standards(conn, wb):
    if "费用标准" not in wb.sheetnames:
        return 0
    ws = wb["费用标准"]
    count = 0
    for row_idx in range(3, ws.max_row + 1):
        grade = text(ws.cell(row_idx, 1).value)
        student_count = int(number(ws.cell(row_idx, 2).value))
        unit_price = ws.cell(row_idx, 3).value
        if not grade or not student_count or unit_price in (None, ""):
            continue
        conn.execute(
            """
            INSERT INTO pricing_standards(grade, student_count, unit_price, description)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(grade, student_count) DO UPDATE SET
              unit_price = excluded.unit_price,
              description = excluded.description
            """,
            (grade, student_count, number(unit_price), text(ws.cell(row_idx, 5).value)),
        )
        count += 1
    return count


def import_teacher_adjustments(conn, wb, month_key, fallback_teacher_names=None, replace=True):
    if "教师薪资汇总" not in wb.sheetnames:
        return 0
    ws = wb["教师薪资汇总"]
    if replace:
        conn.execute("DELETE FROM teacher_adjustments_monthly WHERE month_key = ?", (month_key,))
    fallback_teacher_names = fallback_teacher_names or []
    count = 0
    for row_idx in range(3, ws.max_row + 1):
        teacher_name = valid_teacher_name(ws.cell(row_idx, 1).value)
        fallback_index = row_idx - 3
        if not teacher_name and fallback_index < len(fallback_teacher_names):
            teacher_name = fallback_teacher_names[fallback_index]
        if not teacher_name:
            if not any(number(ws.cell(row_idx, col).value) for col in range(4, 8)):
                continue
            raise RuntimeError(
                f"教师薪资汇总第 {row_idx} 行有车票金额，但无法识别教师姓名。"
                "请先用 Excel/WPS 打开并保存，让 A 列公式计算出老师姓名；"
                "或使用 --teacher-order \"老师1,老师2,...\" 指定薪资汇总行顺序。"
            )
        transports = [number(ws.cell(row_idx, col).value) for col in range(4, 8)]
        notes = text(ws.cell(row_idx, 9).value)
        if not any(transports) and not notes and row_idx - 3 >= len(fallback_teacher_names):
            continue
        upsert_teacher(conn, teacher_name)
        conn.execute(
            """
            INSERT INTO teacher_adjustments_monthly(
              teacher_name, month_key, week1_transport, week2_transport, week3_transport, week4_transport, notes
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(teacher_name, month_key) DO UPDATE SET
              week1_transport = excluded.week1_transport,
              week2_transport = excluded.week2_transport,
              week3_transport = excluded.week3_transport,
              week4_transport = excluded.week4_transport,
              notes = excluded.notes
            """,
            (
                teacher_name,
                month_key,
                transports[0],
                transports[1],
                transports[2],
                transports[3],
                notes,
            ),
        )
        count += 1
    return count


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", help="Excel 工作簿路径")
    parser.add_argument("--month", help="导入到指定月份，格式 YYYY-MM-01；不传时自动从工作簿 Q1 / 文件名 / 总表页名识别")
    parser.add_argument("--db", default=os.path.join(os.path.dirname(__file__), "..", "data", "liming-local.sqlite"))
    parser.add_argument("--append", action="store_true", help="追加课程，不按工作簿中出现的日期替换已有课程")
    parser.add_argument("--no-backup", action="store_true", help="导入前不创建数据库备份")
    parser.add_argument("--skip-teacher-adjustments", action="store_true", help="不导入「教师薪资汇总」页里的车票")
    parser.add_argument(
        "--teacher-order",
        help="当教师薪资汇总 A 列无法读取时，手动指定第 3 行起的老师顺序，例如：何君,李骥,陆俊诚",
    )
    args = parser.parse_args()

    workbook_path = os.path.abspath(args.workbook)
    db_path = os.path.abspath(args.db)
    if not os.path.exists(workbook_path):
        raise FileNotFoundError(workbook_path)

    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=False)
    total_sheet_name = find_total_sheet(wb)

    month_key = infer_month_key(workbook_path, wb, total_sheet_name, args.month)

    conn = sqlite3.connect(db_path, isolation_level=None)
    try:
        conn.execute("PRAGMA busy_timeout = 10000")
        conn.execute("PRAGMA foreign_keys = ON")
        ensure_schema_exists(conn)
        backup_path = "" if args.no_backup else backup_database(conn, db_path)
        conn.execute("BEGIN IMMEDIATE")
        conn.execute("UPDATE settings SET value = ? WHERE key = 'month_key'", (month_key,))
        lessons, lesson_payload = import_lessons(conn, wb, total_sheet_name, month_key, replace=not args.append)
        recharges = import_recharges(conn, wb, month_key, replace=not args.append)
        student_prices = import_student_pricing(conn, wb)
        standards = import_pricing_standards(conn, wb)
        adjustment_teachers = parse_teacher_order(args.teacher_order)
        teacher_adjustments = 0 if args.skip_teacher_adjustments else import_teacher_adjustments(
            conn,
            wb,
            month_key,
            fallback_teacher_names=adjustment_teachers,
            replace=not args.append,
        )
        conn.commit()
    finally:
        conn.close()

    print(f"导入完成：{workbook_path}")
    print(f"月份：{month_key}")
    if backup_path:
        print(f"备份：{backup_path}")
    print(f"课程：{lessons}")
    month_counts = {}
    for row in lesson_payload:
        month_counts[row["month_key"]] = month_counts.get(row["month_key"], 0) + 1
    if month_counts:
        print("课程按实际日期归属：" + "，".join(f"{key}: {value}" for key, value in sorted(month_counts.items())))
    print(f"充值记录：{recharges}")
    print(f"学生单价：{student_prices}")
    print(f"费用标准：{standards}")
    print(f"教师交通费：{teacher_adjustments}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"导入失败：{exc}", file=sys.stderr)
        sys.exit(1)
