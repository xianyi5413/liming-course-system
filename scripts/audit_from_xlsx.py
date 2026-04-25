import argparse
import csv
import datetime as dt
import json
import os
import re
import sqlite3
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.utils.datetime import from_excel


ROOT_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT_DIR / "data"
DB_PATH = DATA_DIR / "liming-local.sqlite"

FIELD_SEVERITY = {
    "teacher_name": "CRITICAL",
    "date": "HIGH",
    "time_slot": "HIGH",
    "classroom": "HIGH",
    "grade": "CRITICAL",
    "subject": "CRITICAL",
    "student_names": "CRITICAL",
    "student_count": "CRITICAL",
    "lesson_status": "MEDIUM",
    "course_status": "MEDIUM",
    "notes": "LOW",
}


def text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def number(value):
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def iso_date(value):
    if value in (None, ""):
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date().isoformat()
        except Exception:
            return ""
    raw = text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return dt.datetime.strptime(raw[:10], fmt).date().isoformat()
        except Exception:
            pass
    return raw


def split_students(value):
    return [part.strip() for part in re.split(r"[、,，;；]", text(value)) if part.strip()]


def canonical_students(value):
    return "、".join(sorted(set(split_students(value))))


def find_total_sheet(workbook, month_key):
    month = int(month_key.split("-")[1])
    preferred = f"{month}月总表"
    if preferred in workbook.sheetnames:
        return preferred
    for name in workbook.sheetnames:
        if re.fullmatch(r"\d{1,2}月总表", name):
            return name
    raise RuntimeError(f"未找到月度总表 Sheet，例如 {preferred}")


def lesson_key(row):
    return (row["date"], row["teacher_name"], row["time_slot"])


def read_xlsx_lessons(xlsx_path, month_key):
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_name = find_total_sheet(workbook, month_key)
    ws = workbook[sheet_name]
    rows = []
    for row_idx in range(3, ws.max_row + 1):
        row = {
            "source_row": row_idx,
            "teacher_name": text(ws.cell(row_idx, 1).value),
            "date": iso_date(ws.cell(row_idx, 2).value),
            "lesson_status": text(ws.cell(row_idx, 3).value),
            "time_slot": text(ws.cell(row_idx, 5).value),
            "classroom": text(ws.cell(row_idx, 6).value),
            "grade": text(ws.cell(row_idx, 7).value),
            "subject": text(ws.cell(row_idx, 8).value),
            "student_names": text(ws.cell(row_idx, 9).value),
            "notes": text(ws.cell(row_idx, 10).value),
            "course_status": text(ws.cell(row_idx, 11).value),
        }
        if not any(row[field] for field in (
            "teacher_name", "date", "lesson_status", "time_slot", "classroom",
            "grade", "subject", "student_names", "notes", "course_status",
        )):
            continue
        row["student_count"] = len(split_students(row["student_names"]))
        rows.append(row)
    return sheet_name, rows


def connect_db(db_path):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    ensure_audit_logs(conn)
    return conn


def ensure_audit_logs(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS audit_logs (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          run_at TEXT DEFAULT CURRENT_TIMESTAMP,
          run_id TEXT DEFAULT '',
          source TEXT,
          severity TEXT,
          entity TEXT,
          field TEXT,
          before_value TEXT,
          after_value TEXT,
          status TEXT DEFAULT 'open',
          notes TEXT DEFAULT ''
        )
        """
    )
    columns = {row["name"] for row in conn.execute("PRAGMA table_info(audit_logs)").fetchall()}
    if "run_id" not in columns:
        conn.execute("ALTER TABLE audit_logs ADD COLUMN run_id TEXT DEFAULT ''")
        conn.commit()


def rows_to_dicts(rows):
    return [dict(row) for row in rows]


def standard_price(conn, grade, student_count):
    if not student_count:
        bucket = 1
    elif str(grade).startswith("高"):
        bucket = 4 if student_count >= 4 else student_count
    elif str(grade).startswith("初"):
        bucket = 4 if student_count >= 3 else min(student_count, 2)
    else:
        bucket = 4 if student_count >= 4 else student_count
    row = conn.execute(
        "SELECT unit_price FROM pricing_standards WHERE grade = ? AND student_count = ?",
        (grade, bucket),
    ).fetchone()
    return number(row["unit_price"]) if row else 0.0


def add_issue(issues, *, source, severity, entity, field, xlsx_value="", db_value="", lesson_id=None, xlsx_row=None, message="", patch=None):
    issue = {
        "source": source,
        "severity": severity,
        "entity": entity,
        "field": field,
        "xlsx_value": "" if xlsx_value is None else str(xlsx_value),
        "db_value": "" if db_value is None else str(db_value),
        "lesson_id": lesson_id,
        "xlsx_row": xlsx_row,
        "message": message,
        "patch": patch or {},
    }
    issues.append(issue)


def compare_lessons(conn, month_key, xlsx_rows):
    issues = []
    db_rows = rows_to_dicts(conn.execute(
        "SELECT * FROM lessons WHERE month_key = ?",
        (month_key,),
    ).fetchall())
    db_by_key = defaultdict(list)
    for row in db_rows:
        db_by_key[(text(row.get("date")), text(row.get("teacher_name")), text(row.get("time_slot")))].append(row)

    matched_db_ids = set()
    xlsx_keys = set()
    for xrow in xlsx_rows:
        key = lesson_key(xrow)
        xlsx_keys.add(key)
        matches = db_by_key.get(key, [])
        if not matches:
            add_issue(
                issues,
                source="xlsx",
                severity="HIGH",
                entity=f"xlsx_row_{xrow['source_row']}",
                field="lesson",
                xlsx_value=f"{xrow['date']} {xrow['teacher_name']} {xrow['time_slot']}",
                db_value="数据库缺失",
                xlsx_row=xrow["source_row"],
                message="xlsx 中存在课程，但数据库未找到 date+teacher+time_slot 完全匹配的记录",
                patch={"type": "insert_lesson", "lesson": xrow},
            )
            continue
        if len(matches) > 1:
            add_issue(
                issues,
                source="xlsx",
                severity="HIGH",
                entity=f"xlsx_row_{xrow['source_row']}",
                field="lesson",
                xlsx_value=f"{xrow['date']} {xrow['teacher_name']} {xrow['time_slot']}",
                db_value=f"匹配到 {len(matches)} 条数据库记录",
                xlsx_row=xrow["source_row"],
                message="数据库存在重复三元组，无法唯一定位课程",
            )
            continue
        db_row = matches[0]
        matched_db_ids.add(db_row["id"])
        lesson_id = db_row["id"]
        for field in ("teacher_name", "date", "lesson_status", "time_slot", "classroom", "grade", "subject", "notes", "course_status"):
            xval = text(xrow.get(field))
            dval = text(db_row.get(field))
            if xval != dval:
                add_issue(
                    issues,
                    source="xlsx",
                    severity=FIELD_SEVERITY[field],
                    entity=f"lesson_{lesson_id}",
                    field=field,
                    xlsx_value=xval,
                    db_value=dval,
                    lesson_id=lesson_id,
                    xlsx_row=xrow["source_row"],
                    patch={"type": "lesson", "id": lesson_id, field: xval},
                )
        xstudents = canonical_students(xrow.get("student_names"))
        dstudents = canonical_students(db_row.get("student_names"))
        if xstudents != dstudents:
            add_issue(
                issues,
                source="xlsx",
                severity=FIELD_SEVERITY["student_names"],
                entity=f"lesson_{lesson_id}",
                field="student_names",
                xlsx_value=xstudents,
                db_value=dstudents,
                lesson_id=lesson_id,
                xlsx_row=xrow["source_row"],
                message="学生名单成员不一致，已按去空白、去重、排序后比较",
                patch={"type": "lesson", "id": lesson_id, "student_names": xrow.get("student_names", "")},
            )
        dcount = len(split_students(db_row.get("student_names")))
        if xrow["student_count"] != dcount:
            add_issue(
                issues,
                source="xlsx",
                severity=FIELD_SEVERITY["student_count"],
                entity=f"lesson_{lesson_id}",
                field="student_count",
                xlsx_value=xrow["student_count"],
                db_value=dcount,
                lesson_id=lesson_id,
                xlsx_row=xrow["source_row"],
                message="学生人数由学生名单拆分后计算",
            )

    for db_row in db_rows:
        key = (text(db_row.get("date")), text(db_row.get("teacher_name")), text(db_row.get("time_slot")))
        if db_row["id"] not in matched_db_ids and key not in xlsx_keys:
            add_issue(
                issues,
                source="xlsx",
                severity="MEDIUM",
                entity=f"lesson_{db_row['id']}",
                field="lesson",
                xlsx_value="xlsx 无对应课程",
                db_value=f"{db_row.get('date', '')} {db_row.get('teacher_name', '')} {db_row.get('time_slot', '')}",
                lesson_id=db_row["id"],
                message="数据库中存在本月课程，但 xlsx 权威源中未找到；可能是事后补录，需要确认",
            )
    return issues


def student_cross_checks(conn, month_key, xlsx_rows):
    issues = []
    by_student = defaultdict(list)
    for row in xlsx_rows:
        for student in split_students(row["student_names"]):
            by_student[student].append((row["source_row"], row["grade"], row["subject"], row["student_count"]))

    students = {row["name"]: dict(row) for row in conn.execute("SELECT * FROM students").fetchall()}
    pricing_rows = rows_to_dicts(conn.execute("SELECT * FROM student_pricing").fetchall())
    pricing_by_student = defaultdict(list)
    for row in pricing_rows:
        pricing_by_student[row["student_name"]].append(row)

    for student, appearances in by_student.items():
        grade_rows = [(row_idx, grade) for row_idx, grade, _, _ in appearances if grade]
        grades = {grade for _, grade in grade_rows}
        if len(grades) > 1:
            add_issue(
                issues,
                source="student_cross",
                severity="CRITICAL",
                entity=f"student_{student}",
                field="grade",
                xlsx_value="; ".join(f"{grade}@row{row_idx}" for row_idx, grade in grade_rows),
                db_value=students.get(student, {}).get("grade", ""),
                message="同一学生在 xlsx 本月课程中出现多个年级",
            )
        latest = next((grade for _, grade in reversed(grade_rows) if grade), "")
        db_grade = text(students.get(student, {}).get("grade", ""))
        if latest and db_grade and latest != db_grade:
            add_issue(
                issues,
                source="student_cross",
                severity="CRITICAL",
                entity=f"student_{student}",
                field="student_grade",
                xlsx_value=latest,
                db_value=db_grade,
                message="students 表年级与 xlsx 最新出现年级不一致",
                patch={"type": "student", "name": student, "grade": latest},
            )
        for pricing in pricing_by_student.get(student, []):
            subject = pricing["subject"]
            matching = [(grade, count) for _, grade, subj, count in appearances if subj == subject and grade]
            if not matching:
                continue
            grade, count = matching[-1]
            std = standard_price(conn, grade, count)
            custom = number(pricing["custom_price"])
            if std > 0 and abs(custom - std) / std > 0.5:
                add_issue(
                    issues,
                    source="student_cross",
                    severity="WARN",
                    entity=f"pricing_{pricing['id']}",
                    field="custom_price",
                    xlsx_value=f"标准价 {std}",
                    db_value=custom,
                    message=f"{student}-{subject} 专享价与按 {grade}/{count} 人计算的标准价差额超过 50%",
                )
    return issues


def write_audit_logs(conn, run_id, issues):
    for issue in issues:
        conn.execute(
            """
            INSERT INTO audit_logs(run_id, source, severity, entity, field, before_value, after_value, status, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'open', ?)
            """,
            (
                run_id,
                issue.get("source", "xlsx"),
                issue.get("severity", ""),
                issue.get("entity", ""),
                issue.get("field", ""),
                issue.get("db_value", ""),
                issue.get("xlsx_value", ""),
                issue.get("message", ""),
            ),
        )
        issue["audit_log_id"] = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.commit()


def write_report_files(report, out_dir):
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M")
    json_path = out_dir / f"audit_report_{stamp}.json"
    csv_path = out_dir / f"audit_report_{stamp}.csv"
    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    with csv_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["审计ID", "严重级别", "来源", "实体", "字段", "xlsx值", "数据库值", "课程ID", "xlsx行号", "说明"])
        for issue in report["issues"]:
            writer.writerow([
                issue.get("audit_log_id", ""),
                issue.get("severity", ""),
                issue.get("source", ""),
                issue.get("entity", ""),
                issue.get("field", ""),
                issue.get("xlsx_value", ""),
                issue.get("db_value", ""),
                issue.get("lesson_id", ""),
                issue.get("xlsx_row", ""),
                issue.get("message", ""),
            ])
    report["report_files"] = {"json": str(json_path), "csv": str(csv_path)}


def run_audit(xlsx_path, month_key, db_path=DB_PATH, write_logs=True, write_files=True):
    run_id = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    sheet_name, xlsx_rows = read_xlsx_lessons(xlsx_path, month_key)
    with connect_db(db_path) as conn:
        issues = compare_lessons(conn, month_key, xlsx_rows)
        issues.extend(student_cross_checks(conn, month_key, xlsx_rows))
        counts = Counter(issue["severity"] for issue in issues)
        report = {
            "run_id": run_id,
            "month_key": month_key,
            "source_file": str(Path(xlsx_path).resolve()),
            "sheet_name": sheet_name,
            "scanned_lessons": len(xlsx_rows),
            "issue_count": len(issues),
            "counts": {key: counts.get(key, 0) for key in ("CRITICAL", "HIGH", "MEDIUM", "LOW", "WARN")},
            "issues": issues,
        }
        if write_logs:
            write_audit_logs(conn, run_id, issues)
        if write_files:
            write_report_files(report, DATA_DIR)
        return report


def main():
    parser = argparse.ArgumentParser(description="从 xlsx 权威源对账课程数据库")
    parser.add_argument("xlsx_path")
    parser.add_argument("--month", required=True)
    parser.add_argument("--db", default=str(DB_PATH))
    parser.add_argument("--stdout-json", action="store_true")
    parser.add_argument("--no-write-files", action="store_true")
    parser.add_argument("--no-write-logs", action="store_true")
    args = parser.parse_args()
    report = run_audit(
        args.xlsx_path,
        args.month,
        db_path=Path(args.db),
        write_logs=not args.no_write_logs,
        write_files=not args.no_write_files,
    )
    if args.stdout_json:
        print(json.dumps(report, ensure_ascii=False))
    else:
        counts = report["counts"]
        print(
            " / ".join(f"{key} {counts.get(key, 0)}" for key in ("CRITICAL", "HIGH", "MEDIUM", "LOW", "WARN"))
        )
        if "report_files" in report:
            print(f"JSON: {report['report_files']['json']}")
            print(f"CSV: {report['report_files']['csv']}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
